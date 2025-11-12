import openpyxl  # type: ignore
from selenium import webdriver  # type: ignore
from selenium.webdriver.common.by import By  # type: ignore
from selenium.webdriver.chrome.options import Options  # type: ignore
from time import sleep

#Janela do Chrome não abrirá
options = Options()
options.add_argument('--headless')
driver = webdriver.Chrome(options=options)


# Carregar planilha de clientes
planilha_clientes = openpyxl.load_workbook(r'AUTOMAÇÕES/Verificar Pagamento/clientes_faturas.xlsx')
pagina_clientes = planilha_clientes['Sheet1']

# Carregar planilha de resultados
planilha_fechamento = openpyxl.load_workbook(r'AUTOMAÇÕES/Verificar Pagamento/pagamento.xlsx')
pagina_fechamento = planilha_fechamento['Sheet1']

# Iniciar WebDriver apenas uma vez
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(1)




# Loop pelos clientes
for i in range(2, pagina_clientes.max_row + 1):
    nome = pagina_clientes.cell(row=i, column=1).value
    cpf = pagina_clientes.cell(row=i, column=2).value
    vencimento = pagina_clientes.cell(row=i, column=3).value

    try:
        campo_pesquisa = driver.find_element(By.XPATH, '//*[@id="cpfInput"]')
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(1)

        driver.find_element(By.XPATH, '//*[@class="btn btn-custom btn-lg btn-block mt-3"]').click()
        sleep(2)

        status = driver.find_element(By.XPATH, '//*[@id="statusLabel"]').text

        if status == 'em dia':
            dataPagamento = driver.find_element(By.XPATH, '//*[@id="paymentDate"]').text
            metodoPagamento = driver.find_element(By.XPATH, '//*[@id="paymentMethod"]').text

            print(f'CPF: {cpf} - Status: {status} - Data Pagamento: {dataPagamento} - Método: {metodoPagamento}')
            pagina_fechamento.append([nome, cpf, status, vencimento, dataPagamento, metodoPagamento])

        else:
            print(f'CPF: {cpf} - Status: {status} (pendente)')
            pagina_fechamento.append([nome, cpf, status, vencimento, 'pendente', '---'])

        sleep(1)

    except Exception as e:
        print(f"Erro ao processar CPF {cpf}: {e}")

# Fechar navegador e salvar resultados
driver.quit()
planilha_fechamento.save(r'AUTOMAÇÕES/Verificar Pagamento/pagamento.xlsx')

print("/n✅ Verificação concluída e resultados salvos com sucesso!")
